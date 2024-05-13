import openpyxl
import random
from openpyxl import Workbook


#deleting some of the rows from master list to create sudo attended list
pathMarch = 'attended.xlsx'
book = openpyxl.load_workbook(pathMarch)
sheet = book.active

randomRemoveRow = random.sample(range(3, sheet.max_row), 20)
print(randomRemoveRow)

for r in randomRemoveRow:
    sheet.delete_rows(r)
book.save('attended.xlsx')
#set up complete

titleRowNum = 2

def findColByName(sheet):
    IDcol = 0
    for c in range(1, sheet.max_column+1):
        if sheet.cell(titleRowNum, c).value.lower() in ["id", "student id"]:
            IDcol = c
            return IDcol
    raise ValueError("Make Sure \"ID\" or \"student id\" Is One Of the Col Title, or Update findColByName Function")



def setIDsDict(path: str, idDict: dict, status: bool):
    book = openpyxl.load_workbook(path)
    sheet = book.active
    maxRow = sheet.max_row
    idCol = findColByName(sheet)

    for r in range(3, maxRow+1):
        id = sheet.cell(r, idCol).value
        idDict[id] = status


def diffIDs(masterPath: str, *minorPaths: str):
    IDdict = {}
    setIDsDict(masterPath, IDdict, False)
    for minor in minorPaths:
        print("hello")
        setIDsDict(minor, IDdict, True)

    absentIDs = []
    for key in IDdict.keys():
        if not IDdict[key]:
            absentIDs.append(key)
    return absentIDs

resultIDs = diffIDs("C3_Nursing_ASN.xlsx", "attended.xlsx")

print(resultIDs, len(resultIDs))